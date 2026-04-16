from datetime import datetime
from itertools import repeat
import openpyxl


def inf_loop(data_loader):
    # wrapper function for endless data loader.
    for loader in repeat(data_loader):
        yield from loader

def save_excel_metadata(metadata_filename, basename, landmarks):
    # Create or load the Excel workbook
    try:
        workbook = openpyxl.load_workbook(metadata_filename)
    except FileNotFoundError:
        # Excel header
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'timestamp'
        sheet['B1'] = 'file'
        idx=1
        count_column=3
        for lm in landmarks:
            value_x = f'x{idx}'
            sheet.cell(row=1, column=count_column, value=value_x)
            count_column = count_column + 1
            value_y = f'y{idx}'
            sheet.cell(row=1, column=count_column, value=value_y)
            count_column = count_column + 1
            value_z = f'z{idx}'
            sheet.cell(row=1, column=count_column, value=value_z)
            count_column = count_column + 1
            idx  = idx + 1
           

    # Select the active sheet
    sheet = workbook.active

    # Find the first empty row in the first column
    row = 1
    while sheet.cell(row=row, column=1).value is not None:
        row += 1

    # Save timestamp & basename in the first column of the first empty row
    timestamp = datetime.now().strftime('%Y-%m-%d-%H:%M:%S')
    time_str = f'{timestamp}'
    sheet.cell(row=row, column=1, value=time_str)
    sheet.cell(row=row, column=2, value=basename)

    # Save landmarks in the same row, starting from the second column
    col = 2
    for lm in landmarks:
        px = lm[0]
        py = lm[1]
        pz = lm[2]
        sheet.cell(row=row, column=col+1, value=px)
        sheet.cell(row=row, column=col+2, value=py)
        sheet.cell(row=row, column=col+3, value=pz)
        col = col + 3

    # Save the workbook
    workbook.save(metadata_filename)


class Timer:
    def __init__(self):
        self.cache = datetime.now()

    def check(self):
        now = datetime.now()
        duration = now - self.cache
        self.cache = now
        return duration.total_seconds()

    def reset(self):
        self.cache = datetime.now()

